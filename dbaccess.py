## スクレイピングした記事をエクセルに出力
import openpyxl as excel

class ScrapingData:
    def __init__(self):
        self.wb = excel.load_workbook('./data.xlsx')
        self.ws = self.wb.active

    def insert_excel(self, today_articles):
        excel_val = []
        row = 2

        # 空行検索、エクセルデータ取得
        while True:
            if self.ws.cell(row=row, column=1).value is None:
                insert_row = row
                break
            else:
                excel_val.append(self.ws.cell(row=row, column=1).value)
                row += 1

        # データ入力
        for val in today_articles:
            column_tx = 'A%d' % insert_row
            column_href = 'B%d' % insert_row
            if val[0] not in excel_val:
                self.ws[column_tx] = val[0]
                self.ws[column_href] = val[1]
                insert_row += 1
        self.wb.save('data.xlsx')
